#!/bin/python3

import re

# excel table
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, Font

import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), "../../../shared/python"))
from cli_parser import parse_args
import shared_constants as cst
from common import get_parts

def get_data(files):
    data = {}

    for file in files:
        # filter paths containing m1, min1, min2, ...
        pattern = re.compile(r'/(min|m)\d+(/|$)')
        if pattern.search(file):
            continue

        if not any(s in file for s in ("opt", "sp")):
            continue

        parts = get_parts(file)
        metal = parts["metal"]
        if metal not in data:
            data[metal] = []

        mod = parts.get("specification", "")
        name = parts["name"]
        element = parts["element"]

        row = next((r for r in data[metal] if r["name"] == name and r["modification"] == mod), None)
        if row is None:
            row = {
                "name": name,
                "element": element,
                "modification": mod,
                "G-E(el) [Eh]": None,
                "E(el) [Eh]": None,
            }
            data[metal].append(row)

        with open(file) as f:
            for l in f:
                l = l.strip()
                if l.startswith("G-E(el)") and parts["c_type"] == "opt":
                    row["G-E(el) [Eh]"] = float(l.split()[2])
                if l.startswith("FINAL SINGLE POINT ENERGY") and parts["c_type"] == "sp":
                    row["E(el) [Eh]"] = float(l.split()[-1])

    return data

def cut(name):
    name = name.replace(".out", "")
    name = name.replace(".inp", "")
    return name

def find_int(name1, name2):
    match = re.match(re.escape(name1) + r"(\d+)", name2)
    number = int(match.group(1)) if match else 1
    return number

def compute_delta_g(data):
    delta_gs = {}
    for metal, rows in data.items():
        if metal == "ligands":
            continue

        precursor = [
            {"name": cut(row["name"]), "G": row["G-E(el) [Eh]"] + row["E(el) [Eh]"],}
            for row in rows
            if any(p in cut(row["name"]) for p in cst.PRECURSOR)
            and not any(v == None for v in (row["G-E(el) [Eh]"], row["E(el) [Eh]"]))
        ]

        if not precursor:
            continue

        for row in rows:
            if any(p in cut(row["name"]) for p in cst.PRECURSOR):
                continue

            if row["E(el) [Eh]"] == None or row["G-E(el) [Eh]"] == None:
                continue

            z_complex = {
                "name": cut(row["name"]), "G": row["G-E(el) [Eh]"] + row["E(el) [Eh]"],
            }

            ligand = [
                {"name": cut(ligand["name"]), "G": ligand["G-E(el) [Eh]"] + ligand["E(el) [Eh]"],}
                for ligand in data["ligands"]
                if cut(ligand["name"]) in cut(row["name"])
                and not any(v == None for v in (ligand["G-E(el) [Eh]"], ligand["E(el) [Eh]"]))
            ]

            precursor_rest = [
                {"name": cut(ligand["name"]), "G": ligand["G-E(el) [Eh]"] + ligand["E(el) [Eh]"],}
                for ligand in data["ligands"]
                if any(cut(ligand["name"]) in cut(p["name"]) for p in precursor)
                and not any(v == None for v in (ligand["G-E(el) [Eh]"], ligand["E(el) [Eh]"]))
            ]

            if any(len(l) != 1 for l in (ligand, precursor_rest, precursor)):
                continue

            pr_mult = find_int(precursor_rest[0]["name"], precursor[0]["name"])
            l_mult = find_int(ligand[0]["name"], z_complex["name"])

            value = (
                    z_complex["G"] + precursor_rest[0]["G"] * pr_mult
                    - ligand[0]["G"] * l_mult - precursor[0]["G"]
            )

            delta_gs[row["name"]] = value*cst.CONSTANTS["Eh->kJ/mol"]

    return delta_gs

def make_sort(data):
    path = "/home/av256/tables/energies.xlsx"

    if os.path.isfile(path):
        mode = "a"
        existing_sheets = load_workbook(path, read_only=True).sheetnames
        sheet_exists = "sorted" in existing_sheets
    else:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        mode = "w"
        sheet_exists = False

    kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a" and sheet_exists:
        kwargs["if_sheet_exists"] = "replace"

    delta_gs = compute_delta_g(data)

    with pd.ExcelWriter(path , **kwargs) as writer:
        first_element, first_rows = next(iter(data.items()))
        first_df = pd.DataFrame(first_rows)

        # Header schreiben
        header_df = pd.DataFrame(columns=first_df.columns)
        header_df.index.name = "#"
        header_df.to_excel(writer, sheet_name="sorted", startrow=0, index=True, header=True)
        ws = writer.sheets["sorted"]

        current_row = 2
        global_index = 1

        max_cols = max(len(pd.DataFrame(rows).columns) for rows in data.values())
        n_cols = max_cols + 2
        dg_col = n_cols + 1

        for element, rows in data.items():
            # Leerzeile mit Element-Label
            ws.cell(row=current_row, column=1, value=element)
            current_row += 1

            df = pd.DataFrame(rows)
            df.index = range(global_index, global_index + len(df))

            for i, (idx, row) in enumerate(df.iterrows()):
                data_row = current_row + i
                ws.cell(row=data_row, column=1, value=idx)
                for j, col in enumerate(df.columns):
                    value = row[col]
                    if isinstance(value, (list, tuple, np.ndarray)):
                        value = ", ".join(str(v) for v in value) if len(value) > 0 else None
                    ws.cell(row=data_row, column=j + 2, value=value)
                    ws.cell(row=data_row, column= 1).font = Font(bold=True)

            # G-Spalte & ΔG-Spalte
            for i in range(len(df)):
                data_row = current_row + i
                
                ws.cell(row=data_row, column=n_cols, value=f"=E{data_row}+F{data_row}")
                
                dg_val = delta_gs.get(rows[i]["name"])
                if dg_val is not None:
                    ws.cell(row=data_row, column=dg_col, value=dg_val)

            global_index += len(df)
            current_row  += len(df)

        # G-Header
        ws.cell(row=1, column=n_cols, value="G [Eh]").font = Font(bold=True)

        # ΔG-Header
        ws.cell(row=1, column=dg_col, value="ΔG [kJ/mol]").font = Font(bold=True)

        # Eine Table über alles
        col_letter = get_column_letter(dg_col)
        total_rows = current_row
        ref = f"A1:{col_letter}{total_rows}"
        table = Table(displayName="TableAll", ref=ref)
        style = TableStyleInfo(name=None, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Rahmen entfernen
        no_border = Border(left=Side(style=None), right=Side(style=None),
                           top=Side(style=None),  bottom=Side(style=None))
        for row in ws.iter_rows(min_row=1, max_row=total_rows, min_col=1, max_col=n_cols):
            for cell in row:
                cell.border = no_border

def create_sort(files):
    data = get_data(files)
    make_sort(data)
