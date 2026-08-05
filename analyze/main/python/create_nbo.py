#!/bin/python3

# excel table
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import Font

import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), "../config"))
import config_constants as cst
sys.path.append(os.path.join(os.path.dirname(__file__), "../../shared/python"))
from cli_parser import parse_args
from common import get_parts

def get_data(files):
    summary_data = []
    nbo_tables = []

    for file in files:
        if "nbo" not in file:
            continue

        parts = get_parts(file)
        row = {
            "metal": parts["metal"],
            "element": parts["element"],
            "ligand": parts["ligand"],
            "modification": parts.get("specification", ""),
            "optional": parts["optional"],
            "name": parts["name"],
        }
        summary_data.append(row)

        table = extract_blocks(file)
        nbo_tables.append(table)

    return summary_data, nbo_tables

def create_nbo_sheet(summary_table, tables):
    path = "/home/av256/tables/energies.xlsx"
    sheet_name = "nbo"

    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    else:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    ws = wb.create_sheet(sheet_name)

    tables_by_name = {t['name']: t for t in tables}
    summary_cols = list(dict.fromkeys(k for entry in summary_table for k in entry.keys()))
    if "name" not in summary_cols:
        raise ValueError("summary_table braucht ein 'name'-Feld für die Verlinkung")

    header_row = 1
    summary_start = header_row + 1
    n_entries = len(summary_table)
    summary_end = summary_start + n_entries - 1

    GAP_AFTER_SUMMARY = 2
    GAP_BETWEEN_ENTRIES = 2
    GAP_BETWEEN_TABLES_COLS = 2

    # --- Pass 1: Zielzeilen berechnen (charge & wbi nebeneinander) ---
    current_row = summary_end + 1 + GAP_AFTER_SUMMARY
    entry_start_rows = []
    for entry in summary_table:
        entry_start_rows.append(current_row)
        t = tables_by_name.get(entry['name'])
        n_charge = len(t['charge']) if t else 0
        n_wbi = len(t['wbi']) if t else 0
        block_len = 1 + 1 + max(n_charge, n_wbi)
        current_row += block_len + GAP_BETWEEN_ENTRIES

    # --- Summary-Header + Daten schreiben ---
    for col_idx, col in enumerate(summary_cols, start=1):
        ws.cell(row=header_row, column=col_idx, value=col).font = Font(bold=True)

    for i, entry in enumerate(summary_table):
        row = summary_start + i
        target = entry_start_rows[i]
        for col_idx, col in enumerate(summary_cols, start=1):
            val = entry.get(col)
            cell = ws.cell(row=row, column=col_idx, value=val if val is not None else "")
            cell.hyperlink = f"#{sheet_name}!A{target}"
        ws.cell(row=row, column=summary_cols.index('name') + 1).style = "Hyperlink"

    # --- Formatierung der Summary-Tabelle (übernommen aus log-Sheet) ---
    for col_idx in range(1, len(summary_cols) + 1):
        col_letter = get_column_letter(col_idx)
        col_cells = ws[col_letter][header_row - 1:summary_end]  # nur Summary-Bereich
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = f"A{summary_start}"

    last_col_letter = get_column_letter(len(summary_cols))
    table_ref = f"A{header_row}:{last_col_letter}{summary_end}"
    tab = Table(displayName="SummaryTable", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=None, showFirstColumn=False, showLastColumn=False,
        showRowStripes=False, showColumnStripes=False,
    )
    ws.add_table(tab)

    no_border = Border(left=Side(style=None), right=Side(style=None),
                        top=Side(style=None), bottom=Side(style=None))
    for row in ws.iter_rows(min_row=header_row, max_row=summary_end,
                             min_col=1, max_col=len(summary_cols)):
        for cell in row:
            cell.border = no_border

    # --- Detailblöcke schreiben ---
    for i, entry in enumerate(summary_table):
        t = tables_by_name.get(entry['name'])
        row = entry_start_rows[i]

        title_cell = ws.cell(row=row, column=1, value=entry['name'])
        title_cell.font = Font(bold=True, size=12)
        title_cell.hyperlink = f"#{sheet_name}!A{summary_start + i}"
        row += 1

        if t is None:
            ws.cell(row=row, column=1, value="Keine Daten gefunden")
            continue

        charge_headers = list(t['charge'][0].keys()) if t['charge'] else []
        for col_idx, h in enumerate(charge_headers, start=1):
            ws.cell(row=row, column=col_idx, value=h).font = Font(bold=True)
        for r_offset, c_entry in enumerate(t['charge'], start=1):
            for col_idx, h in enumerate(charge_headers, start=1):
                ws.cell(row=row + r_offset, column=col_idx, value=c_entry[h])

        wbi_startcol = len(charge_headers) + GAP_BETWEEN_TABLES_COLS + 1
        wbi_headers = list(t['wbi'][0].keys()) if t['wbi'] else []
        for col_offset, h in enumerate(wbi_headers):
            ws.cell(row=row, column=wbi_startcol + col_offset, value=h).font = Font(bold=True)
        for r_offset, w_entry in enumerate(t['wbi'], start=1):
            for col_offset, h in enumerate(wbi_headers):
                ws.cell(row=row + r_offset, column=wbi_startcol + col_offset, value=w_entry[h])

    wb.save(path)

def extract_blocks(file):
    blocks = {
        "charge": {
            "start": "Summary of Natural Population Analysis",
            "end": "* Total *",
            "parser": parse_charge,
            },
        "wbi": {
            "start": "Wiberg bond index matrix in the NAO basis",
            "end": "Wiberg bond index, Totals by atom",
            "parser": parse_wbi,
            },
        "hybrid": {
            "start": "(Occupancy)   Bond orbital / Coefficients / Hybrids",
            "end": "NHO DIRECTIONALITY AND BOND BENDING",
            "parser": parse_hybrid,
            },
        "sopt": {
            "start": "SECOND ORDER PERTURBATION THEORY ANALYSIS OF FOCK MATRIX IN NBO BASIS",
            "end": "NATURAL BOND ORBITALS (Summary)",
            "parser": parse_sopt,
            },
    }

    raw_blocks = {name: [] for name in blocks}

    state = None

    with open(file) as f:
        for l in f:
            l = l.strip()
            matched = False

            for name, info in blocks.items():
                if l.startswith(info["start"]):
                    state = name
                    matched = True
                    break
                if state == name and l.startswith(info["end"]):
                    state = None
                    matched = True
                    break

            if matched:
                continue
            else:
                if not state:
                    continue
                raw_blocks[state].append(l)

    results = {"name": file.split("/")[-1]}

    charges_result, atoms = blocks["charge"]["parser"](raw_blocks["charge"])
    results["charge"] = charges_result

    for name, cfg in blocks.items():
        if name == "charge":
            continue
        if name == "wbi":
            results[name] = cfg["parser"](raw_blocks[name], atoms)
#        else:
#            results[name] = cfg["parser"](raw_blocks[name])

    return results

def parse_charge(blocks):
    charges = []
    atoms = {}
    for l in blocks:
        items = l.split()
        if not items:
            continue
        atom = items[0].lower().rstrip('0123456789')
        if any(atom in lmnt for lmnt in (cst.ELEMENTS, cst.METALS, cst.LIGAND_ELEMENTS)):
            charges.append({
                "atom": f"{items[0]}{items[1]}",
                "natural charge": float(items[2]),
                "core": float(items[3]),
                "valence": float(items[4]),
                "rydberg": float(items[5]),
                "total": float(items[6]),
            })

            atoms[items[1]] = items[0]

    return charges, atoms

def parse_wbi(blocks, atoms):
    pairs = []
    done_pairs = set()

    for l in blocks:
        items = l.split()
        if not items:
            continue
        if items[0] == "Atom":
            items.pop(0)
            matched = {match: i+2 for i, match in enumerate(items) if match in atoms}
            continue

        if not matched:
            continue

        for i in matched.keys():
            j = items[0][:-1]
            if j == i:
                continue
            if j in atoms:
                pair_key = frozenset({j, i})
                if pair_key in done_pairs:
                    continue
                done_pairs.add(pair_key)

                pairs.append({
                    "pair": f"{items[1]}{j}-{atoms[i]}{i}",
                    "WBI": items[matched[i]],
                })

    return pairs

def parse_hybrid(blocks):
    return blocks

def parse_sopt(blocks):
    return blocks

def create_nbo(files):
    summ_data, tables = get_data(files)
#    print(summ_data)
#    for t in tables:
#        print(t["name"])
#        print(t["charge"])
#        print(t["wbi"])
    create_nbo_sheet(summ_data, tables)
    
