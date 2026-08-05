#!/bin/python3

# excel table
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, Font

# extern imports
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), "../../shared/python"))
from common import get_parts
sys.path.append(os.path.join(os.path.dirname(__file__), "../../config"))
import config_constants as cst

def get_data(files):
    data = []

    for file in files:
        parts = get_parts(file)

        row = dict(cst.ROW)
        row["metal"] = parts["metal"]
        row["element"] = parts["element"]
        row["ligand"] = parts["ligand"]
        row["modification"] = parts.get("specification", "")
        row["optional"] = parts["optional"]
        row["c_type"] = parts["c_type"]
        row["name"] = parts["name"]

        with open(file) as f:
            for l in f:
                l = l.strip()
                if l.startswith("FINAL SINGLE POINT ENERGY") and parts["c_type"] in ("opt", "sp"):
                    row["E(el) [Eh]"] = float(l.split()[-1])
                if l.startswith("Final Gibbs free energy") and parts["c_type"] == "opt":
                    row["G [Eh]"] = float(l.split()[-2])
        data.append(row)

    return data

def make_log(data):
    path = "/home/av256/tables/energies.xlsx"

    if os.path.isfile(path):
        mode = "a"
    else:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        mode = "w"
    kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(path , **kwargs) as writer:
        # write data to excel
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="log", index=False)

        # format data to texcel table
        ws = writer.sheets["log"]
        n_rows = ws.max_row
        n_cols = ws.max_column

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"

        ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
        table = Table(displayName="LogTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name=None, showFirstColumn=False,
                                              showLastColumn=False, showRowStripes=False,
                                              showColumnStripes=False)
        ws.add_table(table)

        no_border = Border(left=Side(style=None), right=Side(style=None),
                           top=Side(style=None),  bottom=Side(style=None))
        for row in ws.iter_rows(min_row=1, max_row=n_rows, min_col=1, max_col=n_cols):
            for cell in row:
                cell.border = no_border

def create_log(files):
    data = get_data(files)
    make_log(data)
