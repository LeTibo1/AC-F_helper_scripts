#!/bin/python3

from log import get_row
from get_pending import parse_list
from openpyxl import load_workbook

import os
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), "../../../shared/python"))
from cli_parser import parse_args
from common import get_parts
sys.path.append(os.path.join(os.path.dirname(__file__), "../../config"))
import config_constants as cst

def normalize(v):
    if v == [] or v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    return v

def normalize_name(v):
    if v is None:
        return None
    return v.split(".", 1)[0]

def match_value(key, excel_val, row_val):
    if key == "name":
        return normalize_name(excel_val) == normalize_name(row_val)
    if key == "optional":
        return parse_list(excel_val) == parse_list(row_val)
    return normalize(excel_val) == normalize(row_val)

def update_excel(row, status):
    IGNORE_KEYS = {"cluster", "date", "clean structure", "comment", "col1", "col2", "col3",}
    is_success = False
    path_xlsx = f"{cst.HOME}/tables/log.xlsx"
    wb = load_workbook(path_xlsx)
    ws = wb["log"]

    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    candidates = [
        excel_row for excel_row in ws.iter_rows(min_row=2)
        if all(
            match_value(key, excel_row[header[key] - 1].value, value)
            for key, value in row.items() if key not in IGNORE_KEYS
        )
    ]

    if len(candidates) > 1 and row["date"]:
        candidates = [
            excel_row for excel_row in candidates
            if match_value("date", excel_row[header["date"] - 1].value, row["date"])
        ]

    if len(candidates) > 1:
        print(" ERROR: Too many matches...")
        sys.exit(1)
    if len(candidates) == 0:
        print(" ERROR: No match found...")
        sys.exit(1)

    candidates[0][header["status"] - 1].value = status
    wb.save(path_xlsx)

def main():
    args = parse_args()
    parts = get_parts(args.file)
    row = get_row(parts, None)
    update_excel(row, args.status)

if __name__ == '__main__':
    main()
