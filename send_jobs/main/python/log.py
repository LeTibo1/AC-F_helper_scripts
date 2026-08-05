#!/bin/python3

# excel table
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# extern imports
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), "../../../shared/python"))
from cli_parser import parse_args
from common import get_parts
sys.path.append(os.path.join(os.path.dirname(__file__), "../../config"))
import config_constants as cst

def get_row(parts, cluster):
    row = cst.ROW
    row["metal"] = parts["metal"]
    row["element"] = parts.get("element", None)
    row["ligand"] = parts["ligand"]
    row["modification"] = parts.get("specification", None)
    row["optional"] = parts.get("optional", None)
    row["c_type"] = parts.get("c_type", None)
    row["name"] = parts["name"]
    row["cluster"] = cluster
    row["date"] = parts["date"]
    row["status"] = "pending"

    return row

def add_row(row):
    path = f"{cst.HOME}/tables/log.xlsx"
    wb = load_workbook(path)
    ws = wb["log"]

    row = {k: ", ".join(v) if isinstance(v, list) else v for k, v in row.items()}

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    next_row = ws.max_row + 1
    for col_idx, key in enumerate(headers, 1):
        ws.cell(row=next_row, column=col_idx, value=row.get(key))

    table = ws.tables["LogTable"]
    table.ref = f"A1:{get_column_letter(ws.max_column)}{next_row}"

    wb.save(path)

def main():
    args = parse_args()
    parts = get_parts(args.file)
    row = get_row(parts, args.cluster)
    add_row(row)

if __name__ == '__main__':
    main()
