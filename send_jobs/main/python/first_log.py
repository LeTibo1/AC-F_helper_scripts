#!/bin/python3

from datetime import datetime

# excel table
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

# extern imports
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), "../../../shared/python"))
from cli_parser import parse_args
from common import get_parts
sys.path.append(os.path.join(os.path.dirname(__file__), "../../config"))
import config_constants as cst

def get_data(files):
    data = []

    for file in files:
        parts = get_parts(file)

        if parts["c_type"] == "opt":
            cluster = "justus"
        else:
            cluster = "GG-cluster"

        row = dict(cst.ROW)
        row["metal"] = parts["metal"]
        row["element"] = parts["element"]
        row["ligand"] = parts["ligand"]
        row["modification"] = parts.get("specification", "")
        row["c_type"] = parts["c_type"]
        row["optional"] = parts["optional"]
        row["name"] = parts["name"]
        row["cluster"] = cluster
        row["date"] = parts["date"]
        row["status"] = "complete"

        data.append(row)

    return data

def make_excel(data):
    with pd.ExcelWriter("~/tables/log.xlsx", engine="openpyxl") as writer:
        # write data to excel
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="log", index=False)

        # format data to texcel table
        ws = writer.sheets["log"]
        n_rows = ws.max_row
        n_cols = ws.max_column

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"

        ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
        table = Table(displayName="LogTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name=None, showFirstColumn=False,
                                              showLastColumn=False, showRowStripes=False,
                                              showColumnStripes=False)
        ws.add_table(table)

        MAX_ROWS_BUFFER = 10000
        green_fill  = PatternFill(start_color="D2FFDC", end_color="D2FFDC", fill_type="solid")
        red_fill    = PatternFill(start_color="FFD9D1", end_color="FFD9D1", fill_type="solid")
        orange_fill = PatternFill(start_color="FBE6C1", end_color="FBE6C1", fill_type="solid")

        last_col_letter = get_column_letter(n_cols)
        zeilen_bereich = f"A2:J{MAX_ROWS_BUFFER}"

        ws.conditional_formatting.add(
            zeilen_bereich,
            FormulaRule(formula=['$J2="complete"'], fill=green_fill)
        )
        ws.conditional_formatting.add(
            zeilen_bereich,
            FormulaRule(formula=['$J2="pending"'], fill=orange_fill)
        )
        ws.conditional_formatting.add(
            zeilen_bereich,
            FormulaRule(formula=['$J2="failed"'], fill=red_fill)
        )

        no_border = Border(left=Side(style=None), right=Side(style=None),
                           top=Side(style=None),  bottom=Side(style=None))
        for row in ws.iter_rows(min_row=1, max_row=n_rows, min_col=1, max_col=n_cols):
            for cell in row:
                cell.border = no_border

def main():
    args = parse_args()
    data = get_data(args.files)
    data.sort(key=lambda d: datetime.strptime(d["date"], "%d.%m.%Y"))
    make_excel(data)

if __name__ == '__main__':
    main()
